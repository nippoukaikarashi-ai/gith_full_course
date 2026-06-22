import openpyxl

def save_to_excel(history):
    """判定履歴のリストをエクセルに書き出す"""
    # 1. 新しいエクセルブック（ファイル）を作る
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "判定結果"

    # 2. 見出しを書く
    sheet["A1"] = "回数"
    sheet["B1"] = "判定結果"

    # 3. 履歴を順番に書き込む
    for i, result in enumerate(history, start=1):
        # A列に回数、 B列に結果
        sheet.cell(row=i+1, column=1).value = i
        sheet.cell(row=i+1, column=2).value = result

    # 4.保存する
    filename = "results.xlsx"
    wb.save(filename)
    print(f"エクセルファイル '{filename}'を作成しました！")

# テスト用の履歴データ
my_history = ["10(偶数)","7(奇数)","100(偶数)"]
save_to_excel(my_history)
